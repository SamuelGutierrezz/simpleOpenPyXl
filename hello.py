from openpyxl.workbook import workbook
from openpyxl import load_workbook



wb = load_workbook('YOUR_FILE.xlsx')

ws = wb.active

name = ws['A2'].value

color = ws['B2'].value

print(f'{name} : {color}')



